from openpyxl import load_workbook
import json
import random
import copy
import logging
from collections import Counter

def isfloat(val):
	try:
		_ = float(val)
	except ValueError:
		return False
	except TypeError:
		return False
	return True


class Variant:

	xml_template = open("template.xml").read()
	question_text_template = "<p>Consider the following statement and its proof.</p>\n<hr>\n" \
			"<p><b>Statement:</b> STATEMENT</p>\n" \
			"<p><b>Proof:</b></p>\n<ol>\nSTEPS_TEXT\n</ol>\n<hr>\n" \
			"<p>Is this proof correct? Select below as appropriate. If the proof is incorrect, in your choice also indicate the number of the <i>first</i> step containing a flawed argument.</p>\n"
	answer_fields = [
			["MODEL_ANS1", "FEEDBACK_MODEL", "SCORE_MODEL"],
			["ALT1_ANS1", "FEEDBACK_ALT1", "SCORE_ALT1"],
			["ALT1_ANS2", "FEEDBACK_ALT2", "SCORE_ALT2"]]

	def __init__(self, original, title, question_variables):
		self.original = copy.deepcopy(original)
		self.alteration_ids = []
		self.mistake_ids = []
		self.has_mistake = False
		self.title = title
		self.question_variables = question_variables

	def is_applicable_alteration(self, alteration, issue_warning=False):
		original_copy = copy.deepcopy(self.original)
		for sub in alteration.substitutions:
			key_found = False
			if sub.operation_type in ["replace_step", "replace_statement"]:
				for i, step in enumerate(original_copy):
					if sub.identifier in step and sub.orig in step:
						key_found = True
					original_copy[i] = step.replace(sub.orig, sub.replacement)
			elif sub.operation_type == "replace_variables":
				if sub.orig in self.question_variables: # don't need to check identifier as there's only one question vars field
					key_found = True
			elif sub.operation_type == "remove_step":
				for i, step in enumerate(original_copy):
					if sub.identifier in step:
						key_found = True
						del original_copy[i]
						break
			else:
				logging.warn(f'Invalid operation type "{sub.operation_type}" in alteration "{alteration.id}".')
			if not key_found:
				if issue_warning:
					logging.warn(f'Invalid substitution in alteration "{alteration.id}" because identifier "{sub.identifier}" or orig "{sub.orig}" not found in proof.')
				return False
		return True

	def apply_alteration(self, alteration, sid):
		original_copy = copy.deepcopy(self.original)
		for sub in alteration.substitutions:
			if sub.operation_type in ["replace_step", "replace_statement"]:
				for i, step in enumerate(original_copy):
					if sub.identifier in step and sub.orig in step:
						self.original[i] = step.replace(sub.orig, sub.replacement)
			elif sub.operation_type == "replace_variables":
				self.question_variables = self.question_variables.replace(sub.orig, sub.replacement)
			elif sub.operation_type == "remove_step":
				for i, step in enumerate(original_copy):
					if sub.identifier in step:
						del self.original[i]
						del original_copy[i]
						break
		self.alteration_ids.append(sid)

	def is_applicable_mistake(self, alteration, issue_warning=False):
		if self.has_mistake:
			# Only allow at most one mistake for now
			return False
		original_copy = copy.deepcopy(self.original)
		for sub in alteration.substitutions:
			key_found = False
			if sub.operation_type in ["replace_step", "replace_statement"]:
				for i, step in enumerate(original_copy):
					if sub.identifier in step and sub.orig in step:
						key_found = True
					original_copy[i] = step.replace(sub.orig, sub.replacement)
			elif sub.operation_type == "replace_variables":
				# Don't need to apply this to a copy as the feedback doesn't depend on it
				if sub.orig in self.question_variables:
					key_found = True
			elif sub.operation_type == "remove_step":
				for i, step in enumerate(original_copy):
					if sub.identifier in step:
						key_found = True
						del original_copy[i]
						break
			else:
				logging.warn(f'Invalid operation type "{sub.operation_type}" in alteration "{alteration.id}".')
			if not key_found:
				if issue_warning:
					logging.warn(f'Invalid substitution in alteration "{alteration.id}" because identifier "{sub.identifier}" or orig "{sub.orig}" not found in proof.')
				return False
	
		for key in alteration.feedback.keys():
			key_found = False
			for step in original_copy:
				if key in step:
					key_found = True
			if not key_found:
				if issue_warning:
					logging.warn(f'Invalid feedback in alteration "{alteration.id}" because match "{key}" not found in proof after applying substitutions of this alteration.')
				return False
		return True

	def apply_mistake(self, alteration, mid):
		self.has_mistake = True
		self.answers = alteration.feedback
		self.default_answer = alteration.default_feedback
		original_copy = copy.deepcopy(self.original)
		for sub in alteration.substitutions:
			if sub.operation_type in ["replace_step", "replace_statement"]:
				for i, step in enumerate(original_copy):
					if sub.identifier in step and sub.orig in step:
						self.original[i] = step.replace(sub.orig, sub.replacement)
			elif sub.operation_type == "replace_variables":
				self.question_variables = self.question_variables.replace(sub.orig, sub.replacement)
			elif sub.operation_type == "remove_step":
				for i, step in enumerate(original_copy):
					if sub.identifier in step:
						del self.original[i]
						del original_copy[i]
						break
		self.mistake_ids.append(mid)

	def get_answer_id(self, string):
		for i, step in enumerate(self.original):
			if string in step:
				return i
		raise ValueError(f"Answer '{string}' not found")

	def choice_to_string(self, choice):
		tf = "true" if choice[1] else "false"
		return f'["{choice[0]}", {tf}, "{choice[2]}"]'

	def render(self):
		steps_text = '\n'.join(['<li>' + step + '</li>' for step in self.original[1:]])
		question_text = Variant.question_text_template.replace("STATEMENT", self.original[0]).replace("STEPS_TEXT", steps_text)
		variant_xml = Variant.xml_template.replace("QUESTION_TEXT", question_text)
		variant_xml = variant_xml.replace("QUESTION_NOTE", f'Mistakes: {self.mistake_ids}, Substitutions: {self.alteration_ids}')
		variant_xml = variant_xml.replace("QUESTION_TITLE", f'{self.title} {self.mistake_ids} {self.alteration_ids}')

		choices = []
		for i, step in enumerate(self.original):
			if i == 0:
				choices.append([i, False, "Yes, the proof is correct"])
			else:
				choices.append([i, False, f"No, there is a mistake in step {i}"])
		if self.has_mistake:
			# We only indicate one answer as correct for the purpose of the model answer.
			# Score handling is done via the PRT, and multiple answer can be correct or partially correct.
			has_true_answer = False
			for ans, field in zip(self.answers.items(), Variant.answer_fields):
				aans, afeedback = ans
				fans, ffeedback, fscore = field
				aid = self.get_answer_id(aans)
				variant_xml = variant_xml.replace(fans, f'"{aid}"')
				variant_xml = variant_xml.replace(ffeedback, afeedback.comment)
				variant_xml = variant_xml.replace(fscore, str(afeedback.score))
				if not has_true_answer:
					choices[aid][1] = True
					has_true_answer = True
			variant_xml = variant_xml.replace("FEEDBACK_ELSE", self.default_answer.comment)
			variant_xml = variant_xml.replace("SCORE_ELSE", str(self.default_answer.score))
		else:
			choices[0][1] = True
			variant_xml = variant_xml.replace("MODEL_ANS1", "\"0\"")
			variant_xml = variant_xml.replace("FEEDBACK_MODEL", "Yes, this proof is correct.")
			variant_xml = variant_xml.replace("FEEDBACK_ELSE", "No, this proof in fact is correct.")
			variant_xml = variant_xml.replace("SCORE_ELSE", str(0.0))
		# Even though the remaining scores are unused, assign 0 to them
		# for valid syntax
		variant_xml = variant_xml.replace("SCORE_MODEL", str(1.0))
		for field in Variant.answer_fields:
			fans, ffeedback, fscore = field
			variant_xml = variant_xml.replace(fscore, str(0.0))
		choices_string = "[" + ", ".join([self.choice_to_string(choice) for choice in choices]) + "]"
		self.question_variables += f"\nlist_of_choices: {choices_string};\n"

		variant_xml = variant_xml.replace("QUESTION_VARIABLES", self.question_variables)
		variant_xml = variant_xml.replace("CHOICES_LIST", "list_of_choices")

		return variant_xml

	def __eq__(self, other):
		return self.alteration_ids == other.alteration_ids and self.mistake_ids == other.mistake_ids


class Feedback:
	def __init__(self, variant_id, match, comment, score=0):
		self.variant_id = variant_id
		self.match = match
		self.comment = comment
		self.score = float(score)


class Substitution:
	def __init__(self, variant_id, operation_type, identifier, orig, replacement):
		self.variant_id = variant_id
		self.operation_type = operation_type
		self.identifier = identifier
		self.orig = orig
		self.replacement = replacement
		if not self.identifier:
			self.identifier = self.orig

	def __str__(self):
		return f"Substitution of variant <{self.variant_id}>, type <{self.operation_type}>, identifier <{self.identifier}>, orig <{self.orig}>, replacement <{self.replacement}>"

class Alteration:
	def __init__(self, id, type, status):
		if type not in ['mistake', 'variant']:
			raise ValueError("Values in second column of variants.xlsx must be either variant or mistake.")
		self.type = type
		if status not in ['enabled', 'disabled', '', None]:
			raise ValueError("Values in third column of variants.xlsx must be either enabled or disabled.")
		self.status = status
		self.id = id
		self.substitutions = []
		self.feedback = {}
		self.default_feedback = None

	def is_enabled(self):
		return self.status != 'disabled'

	def is_mistake(self):
		return self.type == 'mistake'

	def add_substitution(self, sub):
		self.substitutions.append(sub)

	def add_feedback(self, k, v):
		self.feedback[k] = v


class VariantGenerator:

	def __init__(self, problem_name):
		wb = load_workbook(filename=f'{problem_name}/proof.xlsx')
		ws = wb.active
		self.title = ws['A1'].value
		self.question_variables = ws['A2'].value or ''
		self.original = [cell.value for cell in ws['A'][2:] if cell.value]

		original_variant = Variant(self.original, self.title, self.question_variables)

		wb = load_workbook(filename=f'{problem_name}/variants.xlsx')
		ws = wb.active
		self.alterations = dict()
		for row in ws.iter_rows(min_row=2):
			alteration = Alteration(row[0].value, row[1].value, row[2].value)
			self.alterations[row[0].value] = alteration

		wb = load_workbook(filename=f'{problem_name}/substitutions.xlsx')
		ws = wb.active
		for row in ws.iter_rows(min_row=2):
			if row[0].value:
				alteration = self.alterations[row[0].value]
				alteration.add_substitution(Substitution(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value))

		wb = load_workbook(filename=f'{problem_name}/feedback.xlsx')
		ws = wb.active
		for row in ws.iter_rows(min_row=2):
			if row[0].value:
				if not isfloat(row[3].value):
					logging.warn(f'Invalid score in feedback row "{[cell.value for cell in row]}"')
					row[3].value = 0
				alteration = self.alterations[row[0].value]
				if not row[1].value:
					alteration.default_feedback = Feedback(row[0].value, '', row[2].value, row[3].value)
				else:
					alteration.add_feedback(row[1].value, Feedback(row[0].value, row[1].value, row[2].value, row[3].value))

		for alteration in self.alterations.values():
			if alteration.type == "mistake":
				original_variant.is_applicable_mistake(alteration, issue_warning=True)
				if not alteration.default_feedback:
					logging.error(f"Alteration {alteration.id} has no default feedback.")
			else:
				original_variant.is_applicable_alteration(alteration, issue_warning=True)

		self.stack_variants = []
		self.mistake_counts = Counter()

	def enabled_alterations(self):
		# Filter iterator?
		return [item for item in self.alterations.items() if item[1].is_enabled()]

	def generate_variants(self, number):
		for _ in range(number):
			status = self.generate_new_variant()
			if not status:
				logging.warn(f"Unable to generate desired number of variants. Generated {len(self.stack_variants)}.")
				return

	def generate_new_variant(self):
		for _ in range(10000):
			variant = self.generate_variant()
			if variant.mistake_ids:
				mistake_id = variant.mistake_ids[0]
			else:
				mistake_id = -1
			if variant not in self.stack_variants and self.mistake_counts[mistake_id] < 10:
				break
		else:
			return False
		logging.info(f"Variant with mistakes {variant.mistake_ids} and alterations {variant.alteration_ids}")
		self.stack_variants.append(variant)
		self.mistake_counts.update([mistake_id])
		return True
		# logging.warn(mistake_id, self.mistake_counts)

	def generate_variant(self):
		variant = Variant(self.original, self.title, self.question_variables)
		for aid, alteration in self.enabled_alterations():
			if not alteration.is_mistake() and variant.is_applicable_alteration(alteration) and random.random() < 0.5:
				variant.apply_alteration(alteration, aid)
		# Permute the list of alterations to make mistake sampling more uniform
		for aid, alteration in random.sample(self.enabled_alterations(), k=len(self.enabled_alterations())):
			if alteration.is_mistake() and variant.is_applicable_mistake(alteration) and random.random() < 0.5:
				variant.apply_mistake(alteration, aid)
				break
		return variant

	def export(self, filename):
		container_data = open("template_container.xml").read()
		container_data = container_data.replace("QUESTIONS_GO_HERE", ''.join([variant.render() for variant in self.stack_variants]))
		open(filename, "w").write(container_data)

logging.basicConfig(filename='error_log.txt', level=logging.INFO, filemode='w')
QUESTION_NAME = 'induction_inequality'
NUMBER_VARIANTS = 30
vg = VariantGenerator(QUESTION_NAME)
vg.generate_variants(NUMBER_VARIANTS)
vg.export(f"{QUESTION_NAME}.xml")
